"""
Export Service for AquaForge
Handles PDF, CSV, and Email export of optimization results.
"""

import pandas as pd
from typing import Dict, List, Any, Optional
from datetime import datetime
import io
import urllib.parse
from swim_ai_reflex.backend.services.base_service import BaseService

try:
    from xhtml2pdf import pisa

    XHTML2PDF_AVAILABLE = True
except ImportError:
    pisa = None
    XHTML2PDF_AVAILABLE = False


class ExportService(BaseService):
    """
    Service to handle data format conversions and export logic.
    """

    def __init__(self):
        super().__init__()

    def _normalize_results(self, results: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Normalize optimization results to flat list of swimmer entries with placement."""
        if not results:
            return []

        # Check if already normalized (has 'team' and 'swimmer' keys)
        first = results[0]
        if "team" in first and "swimmer" in first:
            return results

        # Normalize from event-based structure (OptimizationResult)
        flat_results = []
        for event_res in results:
            event_name = event_res.get("event", "Unknown")
            event_num = event_res.get("event_number", 0)
            is_diving = "diving" in str(event_name).lower()

            # Collect all entries for this event to determine places
            event_entries = []

            # Process Seton
            seton_swimmers = event_res.get("seton_swimmers", [])
            seton_times = event_res.get("seton_times", [])
            seton_points = event_res.get("seton_points", [])

            for i, swimmer in enumerate(seton_swimmers):
                time_val = seton_times[i] if i < len(seton_times) else "0"
                points_val = seton_points[i] if i < len(seton_points) else 0

                # Convert time/score for sorting
                try:
                    sort_val = (
                        float(time_val) if time_val else (0.0 if is_diving else 9999.0)
                    )
                except (ValueError, TypeError):
                    sort_val = 0.0 if is_diving else 9999.0

                event_entries.append(
                    {
                        "event": event_name,
                        "event_num": event_num,
                        "team": "Seton",
                        "swimmer": swimmer,
                        "time": time_val,
                        "points": points_val,
                        "grade": None,
                        "sort_val": sort_val,
                        "is_diving": is_diving,
                    }
                )

            # Process Opponent
            opp_swimmers = event_res.get("opponent_swimmers", [])
            opp_times = event_res.get("opponent_times", [])
            opp_points = event_res.get("opponent_points", [])

            for i, swimmer in enumerate(opp_swimmers):
                time_val = opp_times[i] if i < len(opp_times) else "0"
                points_val = opp_points[i] if i < len(opp_points) else 0

                try:
                    sort_val = (
                        float(time_val) if time_val else (0.0 if is_diving else 9999.0)
                    )
                except (ValueError, TypeError):
                    sort_val = 0.0 if is_diving else 9999.0

                event_entries.append(
                    {
                        "event": event_name,
                        "event_num": event_num,
                        "team": "Opponent",
                        "swimmer": swimmer,
                        "time": time_val,
                        "points": points_val,
                        "grade": None,
                        "sort_val": sort_val,
                        "is_diving": is_diving,
                    }
                )

            # Sort entries
            # Diving: Descending score (sort_val)
            # Swimming: Ascending time (sort_val)
            event_entries.sort(key=lambda x: x["sort_val"], reverse=is_diving)

            # Assign places and add to flat results
            for place, entry in enumerate(event_entries, 1):
                entry["place"] = place
                del entry["sort_val"]  # Cleanup temp field
                del entry["is_diving"]
                flat_results.append(entry)

        return flat_results

    def to_pdf(self, html_content: str) -> bytes:
        """
        Convert HTML content to PDF bytes.
        """
        if not XHTML2PDF_AVAILABLE:
            self.log_warning("xhtml2pdf not installed, PDF generation unavailable")
            return b""

        try:
            pdf_buffer = io.BytesIO()
            pisa_status = pisa.CreatePDF(io.StringIO(html_content), dest=pdf_buffer)

            if pisa_status.err:
                self.log_error(f"PDF generation error: {pisa_status.err}")
                return b""

            return pdf_buffer.getvalue()
        except Exception as e:
            self.log_error(f"PDF generation exception: {str(e)}")
            return b""

    def to_csv(
        self,
        optimization_results: List[Dict[str, Any]],
        seton_score: float,
        opponent_score: float,
        metadata: Optional[Dict[str, str]] = None,
    ) -> str:
        """
        Convert optimization results to CSV format.
        """
        if not optimization_results:
            self.log_warning("Exporting empty CSV results")
            return "No optimization results available"

        # Normalize data
        optimization_results = self._normalize_results(optimization_results)

        try:
            # Convert to DataFrame
            df = pd.DataFrame(optimization_results)

            # Create header with metadata
            header_lines = [
                "# AquaForge Optimization Results",
                f"# Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                f"# Seton Score: {seton_score}",
                f"# Opponent Score: {opponent_score}",
            ]

            if metadata:
                for key, value in metadata.items():
                    header_lines.append(f"# {key}: {value}")

            header_lines.append("")  # Blank line before data

            # Convert DataFrame to CSV
            csv_buffer = io.StringIO()
            df.to_csv(csv_buffer, index=False)
            csv_data = csv_buffer.getvalue()

            self.log_info("Generated CSV export")
            return "\n".join(header_lines) + "\n" + csv_data

        except Exception as e:
            self.log_error(f"CSV generation failed: {str(e)}")
            return ""

    def to_html_table(
        self,
        optimization_results: List[Dict[str, Any]],
        seton_score: float,
        opponent_score: float,
    ) -> str:
        """
        Convert optimization results to HTML table.
        Used for email and PDF generation.
        """
        if not optimization_results:
            return "<p>No optimization results available</p>"

        # Normalize data
        optimization_results = self._normalize_results(optimization_results)

        # Validate normalized data
        if not optimization_results:
            self.log_warning("Normalized optimization results are empty")
            return "<p>No valid data to export</p>"

        first_row = optimization_results[0]
        required_keys = ["team", "swimmer", "event", "time", "points"]
        missing = [k for k in required_keys if k not in first_row]
        if missing:
            self.log_error(f"Export data missing keys: {missing}")
            return f"<p>Data validation failed: Missing fields {missing}</p>"

        try:
            # Start HTML
            html = f"""
            <html>
            <head>
                <style>
                    :root {{
                        --bg-color: #0a0f1d;
                        --text-color: white;
                        --sub-text: rgba(255,255,255,0.5);
                        --card-bg: rgba(255,255,255,0.08);
                        --header-bg: rgba(197,160,40,0.15);
                        --header-text: #E6D28C;
                        --header-border: rgba(197,160,40,0.3);
                        --border-color: rgba(255,255,255,0.05);
                        --opponent-bg: rgba(255,255,255,0.03);
                        --accent: #E6D28C;
                        --points-color: #4ade80;
                        --time-color: rgba(255,255,255,0.9);
                    }}
                    body.light-mode {{
                        --bg-color: #ffffff;
                        --text-color: #111111;
                        --sub-text: #666666;
                        --card-bg: #f3f4f6;
                        --header-bg: #fef9c3;
                        --header-text: #854d0e;
                        --header-border: #fde047;
                        --border-color: #e5e7eb;
                        --opponent-bg: #f9fafb;
                        --accent: #C5A028;
                        --points-color: #15803d;
                        --time-color: #374151;
                    }}
                    @media print {{
                        body {{ 
                            --bg-color: white !important;
                            --text-color: black !important;
                            --sub-text: #444 !important;
                            --card-bg: #fff !important;
                            --header-bg: #eee !important;
                            --header-text: black !important;
                            --header-border: #ccc !important;
                            --border-color: #ddd !important;
                            --opponent-bg: #fff !important;
                            --accent: black !important;
                            --points-color: black !important;
                            --time-color: black !important;
                        }}
                        .print-toggle {{ display: none !important; }}
                    }}

                    body {{ font-family: 'Inter', Arial, sans-serif; background-color: var(--bg-color); color: var(--text-color); padding: 40px; transition: background-color 0.3s, color 0.3s; }}
                    h1 {{ color: var(--accent); text-align: center; margin-bottom: 5px; font-weight: 300; letter-spacing: 2px; text-transform: uppercase; }}
                    .subtitle {{ text-align: center; color: var(--sub-text); font-size: 0.9em; margin-bottom: 40px; }}
                    
                    /* Toggle Button */
                    .print-toggle {{ position: fixed; top: 20px; right: 20px; padding: 10px 20px; background: var(--card-bg); color: var(--text-color); border: 1px solid var(--border-color); border-radius: 6px; cursor: pointer; font-family: inherit; font-size: 0.85em; transition: all 0.2s; }}
                    .print-toggle:hover {{ background: var(--border-color); }}

                    /* Score Table */
                    .score-container {{ width: 100%; max-width: 600px; margin: 0 auto 40px auto; }}
                    .score-table {{ width: 100%; border-collapse: separate; border-spacing: 20px 0; }}
                    .score-card {{ background: var(--card-bg); border-radius: 8px; border-top: 4px solid var(--accent); padding: 20px; text-align: center; box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1); }}
                    .score-val {{ font-size: 3em; font-weight: 700; color: var(--text-color); display: block; margin: 10px 0; }}
                    .score-lbl {{ color: var(--accent); font-size: 0.85em; letter-spacing: 2px; text-transform: uppercase; }}
                    
                    /* Main Table */
                    .main-table {{ width: 100%; max-width: 900px; margin: 0 auto; border-collapse: collapse; color: var(--text-color); font-size: 0.95em; }}
                    .evt-header {{ background: var(--header-bg); color: var(--header-text); font-weight: 700; padding: 12px 20px; text-align: left; border-top: 1px solid var(--header-border); font-size: 1.1em; }}
                    .team-header {{ background: var(--opponent-bg); padding: 10px 20px; font-weight: 600; color: var(--sub-text); font-size: 0.9em; text-transform: uppercase; letter-spacing: 1px; }}
                    .col-header {{ color: var(--sub-text); font-size: 0.75em; border-bottom: 1px solid var(--border-color); }}
                    .col-header td {{ padding: 5px 0 5px 0; }}
                    
                    .data-row td {{ padding: 8px 0; border-bottom: 1px solid var(--border-color); }}
                    .opponent-row td {{ background: var(--opponent-bg); }}
                    .data-row:last-child td {{ border-bottom: none; }}
                    .swimmer-name {{ padding-left: 30px; }}
                    .time-val {{ font-family: 'Courier New', monospace; color: var(--time-color); }}
                    .points-val {{ font-weight: 700; color: var(--points-color); }}
                    
                    .footer {{ margin-top: 50px; text-align: center; color: var(--sub-text); font-size: 0.8em; border-top: 1px solid var(--border-color); padding-top: 20px; }}
                </style>
                <script>
                    function toggleTheme() {{
                        document.body.classList.toggle('light-mode');
                        const btn = document.querySelector('.print-toggle');
                        if (document.body.classList.contains('light-mode')) {{
                            btn.textContent = 'Switch to Dark Mode';
                        }} else {{
                            btn.textContent = 'Print Friendly View';
                        }}
                    }}
                </script>
            </head>
            <body>
                <button class="print-toggle" onclick="toggleTheme()">Print Friendly View</button>
                <h1>AquaForge Analysis</h1>
                <div class="subtitle">Generated by Admiral Koehr Computation Model</div>
                
                <div class="score-container">
                    <table class="score-table">
                        <tr>
                            <td class="score-card">
                                <span class="score-lbl">Seton Swimming (Projected)</span>
                                <span class="score-val">{seton_score}</span>
                            </td>
                            <td class="score-card">
                                <span class="score-lbl">Opponent</span>
                                <span class="score-val">{opponent_score}</span>
                            </td>
                        </tr>
                    </table>
                </div>
                
                <table class="main-table">
                    <tbody>
            """

            # Group by Event
            df_results = pd.DataFrame(optimization_results)

            # Define standard event order for sorting
            event_order = [
                "200 Medley Relay",
                "200 Yard Medley Relay",
                "Boys 200 Yard Medley Relay",
                "Girls 200 Yard Medley Relay",
                "200 Free",
                "200 Yard Freestyle",
                "Boys 200 Yard Freestyle",
                "Girls 200 Yard Freestyle",
                "200 IM",
                "200 Yard IM",
                "Boys 200 Yard IM",
                "Girls 200 Yard IM",
                "50 Free",
                "50 Yard Freestyle",
                "Boys 50 Yard Freestyle",
                "Girls 50 Yard Freestyle",
                "Diving",
                "1 Meter Diving",
                "Boys 1 Meter Diving",
                "Girls 1 Meter Diving",
                "100 Fly",
                "100 Yard Butterfly",
                "Boys 100 Yard Butterfly",
                "Girls 100 Yard Butterfly",
                "100 Free",
                "100 Yard Freestyle",
                "Boys 100 Yard Freestyle",
                "Girls 100 Yard Freestyle",
                "500 Free",
                "500 Yard Freestyle",
                "Boys 500 Yard Freestyle",
                "Girls 500 Yard Freestyle",
                "200 Free Relay",
                "200 Yard Freestyle Relay",
                "Boys 200 Yard Freestyle Relay",
                "Girls 200 Yard Freestyle Relay",
                "100 Back",
                "100 Yard Backstroke",
                "Boys 100 Yard Backstroke",
                "Girls 100 Yard Backstroke",
                "100 Breast",
                "100 Yard Breaststroke",
                "Boys 100 Yard Breaststroke",
                "Girls 100 Yard Breaststroke",
                "400 Free Relay",
                "400 Yard Freestyle Relay",
                "Boys 400 Yard Freestyle Relay",
                "Girls 400 Yard Freestyle Relay",
            ]

            # Helper to get sort index
            def get_event_index(evt_name):
                # Try exact match first
                if evt_name in event_order:
                    return event_order.index(evt_name)
                # Try partial match
                for i, known in enumerate(event_order):
                    if known in evt_name:
                        return i
                return 999

            # Sort events
            unique_events = sorted(df_results["event"].unique(), key=get_event_index)

            for event in unique_events:
                event_rows = df_results[df_results["event"] == event]

                # Calculate calculated event scores
                # Ensure points are numeric
                seton_evt_score = (
                    event_rows[event_rows["team"] == "Seton"]["points"]
                    .apply(lambda x: float(x) if x is not None else 0)
                    .sum()
                )
                opp_evt_score = (
                    event_rows[event_rows["team"] == "Opponent"]["points"]
                    .apply(lambda x: float(x) if x is not None else 0)
                    .sum()
                )

                html += f"""
                <tr>
                    <td colspan="3" class="evt-header">
                        {event}
                        <span style="float:right; font-size: 0.8em; opacity: 0.7; font-weight: normal;">
                            Seton {seton_evt_score:.0f} - {opp_evt_score:.0f} Opponent
                        </span>
                    </td>
                </tr>
                """

                # Group by Team
                # Sort teams: Seton first
                teams = sorted(
                    event_rows["team"].unique(),
                    key=lambda t: 0 if "seton" in t.lower() else 1,
                )

                for team in teams:
                    team_rows = event_rows[event_rows["team"] == team]
                    # Sort by points desc, then time asc
                    team_rows = team_rows.sort_values(
                        ["points", "time"], ascending=[False, True]
                    )

                    # Team Header Row
                    html += f"""
                    <tr>
                        <td colspan="3" class="team-header">{team}</td>
                    </tr>
                    <tr class="col-header">
                        <td style="padding-left: 30px;">SWIMMER</td>
                        <td>TIME</td>
                        <td>SCORE</td>
                    </tr>
                    """

                    for _, row in team_rows.iterrows():
                        swimmer = row.get("swimmer", "N/A")
                        time_raw = row.get("time", 0)
                        points = row.get("points", 0)

                        # Format time to 2 decimals
                        try:
                            time_secs = float(time_raw)
                            # Format MM:SS
                            if time_secs >= 9999:
                                time_mmss = "--:--"
                            elif time_secs >= 60:
                                mins = int(time_secs // 60)
                                secs = time_secs % 60
                                time_mmss = f"{mins}:{secs:05.2f}"
                            else:
                                time_mmss = f"{time_secs:.2f}"
                        except (ValueError, TypeError):
                            time_mmss = str(time_raw)

                        # Format points to 2 decimals
                        try:
                            pts_display = f"{float(points):.2f}"
                        except (ValueError, TypeError):
                            pts_display = str(points)

                        # Style points - color override if needed (base class handles weight)
                        # .points-val handles green color. Only override if 0
                        pts_color_style = ""
                        pts_color_style = ""
                        if points <= 0:
                            pts_color_style = 'style="color: rgba(255,255,255,0.3);"'

                        row_class = "data-row"
                        if "opponent" in team.lower():
                            row_class += " opponent-row"

                        html += f"""
                        <tr class="{row_class}">
                            <td class="swimmer-name">{swimmer}</td>
                            <td class="time-val">{time_mmss}</td>
                            <td class="points-val" {pts_color_style}>{pts_display}</td>
                        </tr>
                        """

            html += """
                    </tbody>
                </table>
            """

            html += f"""
                <div class="footer">
                    <p>Powered by AquaForge • Generated {datetime.now().strftime("%B %d, %Y at %I:%M %p")}</p>
                </div>
            </body>
            </html>
            """

            self.log_info("Generated HTML table export")
            return html

        except Exception as e:
            self.log_error(f"HTML table generation failed: {str(e)}")
            return "<p>Error generating report</p>"

    def prepare_email_body(
        self,
        optimization_results: List[Dict[str, Any]],
        seton_score: float,
        opponent_score: float,
        recipient_name: str = "Coach",
    ) -> str:
        """
        Prepare plain text email body.
        """
        # Normalize data
        optimization_results = self._normalize_results(optimization_results)

        try:
            lines = [
                f"Hello {recipient_name},",
                "",
                "Here are the optimization results from AquaForge:",
                "",
                f"Seton Swimming: {seton_score} points",
                f"Opponent: {opponent_score} points",
                f"Projected Margin: {'+' if seton_score > opponent_score else ''}{seton_score - opponent_score:.1f} points",
                "",
                "Optimized Lineup:",
                "-" * 70,
            ]

            # Add lineup details
            for result in optimization_results:
                event = result.get("event", "N/A")
                swimmer = result.get("swimmer", "N/A")
                time = result.get("time", "N/A")
                points = result.get("points", "N/A")

                lines.append(
                    f"{event:30} | {swimmer:20} | {str(time):8} | {points} pts"
                )

            lines.extend(
                [
                    "-" * 70,
                    "",
                    "Generated by Admiral Koehr Computation Model",
                    f"{datetime.now().strftime('%B %d, %Y at %I:%M %p')}",
                    "",
                    "Best of luck at the meet!",
                    "",
                    "- AquaForge",
                ]
            )

            self.log_info("Generated email body")
            return "\n".join(lines)

        except Exception as e:
            self.log_error(f"Email body generation failed: {str(e)}")
            return "Error generating email."

    def to_xlsx(
        self,
        optimization_results: List[Dict[str, Any]],
        seton_score: float,
        opponent_score: float,
        seton_team_name: str = "Seton Swimming",
        opponent_team_name: str = "Opponent",
    ) -> bytes:
        """
        Convert optimization results to Excel (.xlsx) format.
        Creates a clean, print-ready lineup for Coach's clipboard.
        """
        if not optimization_results:
            self.log_warning("Exporting empty Excel results")
            return b""

        # Normalize data
        optimization_results = self._normalize_results(optimization_results)

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Meet Lineup"

            # Define styles
            header_font = Font(bold=True, size=14, color="FFFFFF")
            header_fill = PatternFill(
                start_color="1a365d", end_color="1a365d", fill_type="solid"
            )  # Navy
            PatternFill(
                start_color="C5A028", end_color="C5A028", fill_type="solid"
            )  # Gold
            event_font = Font(bold=True, size=11)
            event_fill = PatternFill(
                start_color="E6D28C", end_color="E6D28C", fill_type="solid"
            )  # Light gold
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # Title Row
            ws.merge_cells("A1:F1")
            ws["A1"] = f"AQUAFORGE LINEUP - {seton_team_name} vs {opponent_team_name}"
            ws["A1"].font = Font(bold=True, size=16, color="1a365d")
            ws["A1"].alignment = Alignment(horizontal="center")

            # Score Row
            ws.merge_cells("A2:C2")
            ws["A2"] = f"{seton_team_name}: {seton_score}"
            ws["A2"].font = Font(bold=True, size=12)
            ws.merge_cells("D2:F2")
            ws["D2"] = f"{opponent_team_name}: {opponent_score}"
            ws["D2"].font = Font(bold=True, size=12)

            # Date Row
            ws.merge_cells("A3:F3")
            ws["A3"] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
            ws["A3"].font = Font(italic=True, size=10, color="666666")

            # Blank row
            row_num = 5

            # Column Headers
            headers = ["Event #", "Event", "Swimmer", "Grade", "Seed Time", "Notes"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

            row_num += 1

            # Sort results by event order
            event_order = [
                "200 Medley Relay",
                "200 Yard Medley Relay",
                "200 Free",
                "200 Yard Freestyle",
                "200 IM",
                "200 Yard IM",
                "50 Free",
                "50 Yard Freestyle",
                "Diving",
                "1 Meter Diving",
                "100 Fly",
                "100 Yard Butterfly",
                "100 Free",
                "100 Yard Freestyle",
                "500 Free",
                "500 Yard Freestyle",
                "200 Free Relay",
                "200 Yard Freestyle Relay",
                "100 Back",
                "100 Yard Backstroke",
                "100 Breast",
                "100 Yard Breaststroke",
                "400 Free Relay",
                "400 Yard Freestyle Relay",
            ]

            def get_event_index(evt):
                evt_lower = str(evt).lower()
                for i, e in enumerate(event_order):
                    if e.lower() in evt_lower or evt_lower in e.lower():
                        return i
                return 999

            # Filter to only Seton swimmers (Coach Jim only needs his team)
            seton_results = [
                r
                for r in optimization_results
                if "seton" in str(r.get("team", "")).lower()
            ]

            # Sort by event
            sorted_results = sorted(
                seton_results,
                key=lambda x: (
                    get_event_index(x.get("event", "")),
                    x.get("time", 9999),
                ),
            )

            # Group by event
            current_event = None
            event_num = 0

            for result in sorted_results:
                event = result.get("event", "Unknown")
                swimmer = result.get("swimmer", "Unknown")
                grade = result.get("grade", "")
                time_val = result.get("time", 0)

                # Format time to MM:SS.ss
                try:
                    secs = float(time_val)
                    if secs >= 60:
                        mins = int(secs // 60)
                        rem = secs % 60
                        time_str = f"{mins}:{rem:05.2f}"
                    else:
                        time_str = f"{secs:.2f}"
                except Exception:
                    time_str = str(time_val)

                # New event = new section
                if event != current_event:
                    current_event = event
                    event_num += 1

                    # Event header row
                    ws.merge_cells(
                        start_row=row_num, start_column=1, end_row=row_num, end_column=6
                    )
                    cell = ws.cell(
                        row=row_num, column=1, value=f"Event {event_num}: {event}"
                    )
                    cell.font = event_font
                    cell.fill = event_fill
                    cell.alignment = Alignment(horizontal="left")
                    row_num += 1

                # Data row
                ws.cell(row=row_num, column=1, value=event_num)
                ws.cell(row=row_num, column=2, value=event)
                ws.cell(row=row_num, column=3, value=swimmer)
                ws.cell(row=row_num, column=4, value=grade if grade else "")
                ws.cell(row=row_num, column=5, value=time_str)
                ws.cell(row=row_num, column=6, value="")  # Empty notes column for coach

                # Style the row
                for col in range(1, 7):
                    ws.cell(row=row_num, column=col).border = thin_border
                    ws.cell(row=row_num, column=col).alignment = Alignment(
                        horizontal="center" if col in [1, 4] else "left"
                    )

                row_num += 1

            # Set column widths
            ws.column_dimensions["A"].width = 10
            ws.column_dimensions["B"].width = 25
            ws.column_dimensions["C"].width = 25
            ws.column_dimensions["D"].width = 8
            ws.column_dimensions["E"].width = 12
            ws.column_dimensions["F"].width = 20

            # Set print area and orientation
            ws.print_title_rows = "5:5"  # Repeat header row on each page
            ws.page_setup.orientation = "landscape"
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0

            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            self.log_info("Generated Excel export for Coach's clipboard")
            return output.getvalue()

        except Exception as e:
            self.log_error(f"Excel generation failed: {str(e)}")
            return b""

    def create_mailto_link(self, to_email: str, subject: str, body: str) -> str:
        """
        Create a mailto URL for email clients.
        """
        try:
            # URL encode the body
            encoded_body = urllib.parse.quote(body)
            encoded_subject = urllib.parse.quote(subject)
            return f"mailto:{to_email}?subject={encoded_subject}&body={encoded_body}"
        except Exception as e:
            self.log_error(f"Mailto link creation failed: {str(e)}")
            return f"mailto:{to_email}"


# Singleton instance
export_service = ExportService()
